import base64
import ctypes
import cv2
import numpy as np
import os
import threading
import time
from os import makedirs
from os.path import exists
import pytesseract
from re import match, M, I
import shutil
import sys
import urllib.error
from urllib.request import urlretrieve
import winreg

import globalValue as Pb
from PyQt5 import QtWidgets
from PyQt5.Qt import Qt
from PyQt5.QtCore import QDate, QSettings
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtWidgets import QFileDialog, QLineEdit, QApplication, QMainWindow, QAction, qApp, QLabel, QPushButton, \
    QHBoxLayout, QVBoxLayout, QWidget, QTableWidget, QTextBrowser, QAbstractItemView, QTableWidgetItem, QMessageBox, \
    QRadioButton, QDateEdit, QComboBox, QSplashScreen
from openpyxl import load_workbook

from setting import ChildWindow

from images.setting_png import img as setting_png
from images.icon_png import img as icon_png
from images.exit_png import img as exit_png
from images.info_png import img as info_png


def load_sheet(file):
    """
    加载Excel表格文件
    :param file: 表格文件
    :return: 表格内的页
    """
    excel_file = load_workbook(file)
    sheet = excel_file[excel_file.sheetnames[0]]  # 切换到第一个工作表
    return sheet


def get_file_list(sheet):
    """
    获取表格内容
    :param sheet: 表格内的页
    :return: res
    """
    res = dict()
    for rows in range(2, 34):  # 4 * 8 = 32  + 2 = 34
        room = sheet.cell(rows, 5).value
        if room is not None:
            dic = {"id": sheet.cell(rows, 4).value, "name": sheet.cell(rows, 3).value,
                   "img": sheet.cell(rows, 6).hyperlink.target}
            if room in res.keys():
                res[room].append(dic)
            else:
                res[room] = [dic]
        else:
            break
    return res


def find_contours(image):
    # 预处理操作
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) #二值化
    gray = cv2.GaussianBlur(gray, (5, 5), 0) #高斯滤波
    edged = cv2.Canny(gray, 75, 200) #canny检测
    # 轮廓检测
    h = cv2.findContours(edged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)  # 寻找轮廓
    contours = h[1] #获取轮廓坐标点
    # 对一系列轮廓点坐标按它们围成的区域面积进行排序，取前5个
    contours = sorted(contours, key=cv2.contourArea, reverse=True)[:5]
    #找到第一个轮廓线为矩形的，即为我们要找的轮廓线
    for contour in contours:
        peri = cv2.arcLength(contour, True)  # 周长，闭合
        approx = cv2.approxPolyDP(contour, 0.02 * peri, True)  # 检测出来的轮廓可能是离散的点，故因在此做近似计算，使其形成一个矩形
        # 做精度控制，原始轮廓到近似轮廓的最大的距离，较小时可能为多边形；较大时可能为矩形
        # True表示闭合
        if len(approx) == 4:  # 如果检测出来的是矩形，则break本段if
            screenCnt = approx
            break
    img = cv2.drawContours(image, [screenCnt], -1, (0, 0, 255), 2)  # 绘制轮廓，-1表示全部绘制
    cv2.imshow("img", img)
    cv2.imwrite('./test_contours.jpg', img) #保存结果图片
    cv2.waitKey(0)
    cv2.destroyAllWindows()
    return screenCnt


class MySplashScreen(QSplashScreen):
    # 鼠标点击事件
    def mousePressEvent(self, event):
        pass


class MainWindow(QMainWindow):
    def __init__(self):
        """
        初始化窗口和组件
        """
        super().__init__()
        self.child_window = None

        self.dir_label = QLabel("保存路径")
        self.file_label = QLabel("文件路径")
        self.apart_label = QLabel("楼栋编号")

        self.dir_edit = QLineEdit()
        self.file_edit = QLineEdit()
        self.apart_edit = QLineEdit()

        self.floor_edit = QComboBox(self)
        self.floor_edit.addItems(['选择楼层', '1楼', '2楼', '3楼', '4楼', '5楼', '6楼'])
        self.time_edit = QComboBox(self)
        self.time_edit.addItems(['不加时间', '早上', '中午', '下午', '晚上'])

        self.date_edit = QDateEdit(self)
        self.date_edit.setDisplayFormat('yyyy年MM月dd日')
        self.date_edit.setMinimumDate(QDate.currentDate().addDays(-3652))
        self.date_edit.setMaximumDate(QDate.currentDate().addDays(3652))
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)

        self.dir_button = QPushButton("选择路径")
        self.file_button = QPushButton("选择文件")
        self.push_button = QPushButton("开始拉取")

        self.radio_button_ky = QRadioButton("抗原检测结果", self)
        self.radio_button_hs = QRadioButton("核酸检测结果", self)
        self.radio_button_hs.setChecked(True)

        self.process_output = QTextBrowser()

        self.cursor = self.process_output.textCursor()

        self.table_widget = QTableWidget()

        self.settings = QSettings("config.ini", QSettings.IniFormat)

        self.child_window = ChildWindow()

        self.init_gui()

    def init_gui(self):
        """
        初始化MainWindow
        :return: null
        """
        self.setFixedSize(960, 540)
        self.setWindowTitle("豪杰收集助手")

        tmp = open("icon.png", "wb")
        tmp.write(base64.b64decode(icon_png))
        tmp.close()
        self.setWindowIcon(QIcon(r"icon.png"))
        os.remove("icon.png")

        self.init_menubar()
        self.init_layout()
        self.add_affairs()

        self.setup()

        splash = MySplashScreen()
        splash.setPixmap(QPixmap('./images/starter.png'))
        splash.show()
        time.sleep(3)
        splash.finish(self)
        splash.deleteLater()

        self.printf(">> 已启用配置文件设置\n")

    def init_menubar(self):
        """
        初始化菜单栏
        :return: null
        """
        menubar = self.menuBar()

        menubar.setStyleSheet('''
            QMenu {background-color:#ffffff;}\
            QMenu::item {color:#000000;background:#ffffff;}\
            QMenu::item:selected {background:#f1f1f1;}\
            QMenu::item:pressed {background:#eeeeee;}";
        ''')

        file_menu = menubar.addMenu("文件(&F)")
        help_menu = menubar.addMenu("帮助(&H)")

        tmp = open("setting.png", "wb")
        tmp.write(base64.b64decode(setting_png))
        tmp = open("exit.png", "wb")  # 创建临时的文件
        tmp.write(base64.b64decode(exit_png))  # 把这个one图片解码出来，写入文件中去。
        tmp = open("info.png", "wb")
        tmp.write(base64.b64decode(info_png))
        tmp.close()

        setting_action = QAction(QIcon(r"setting.png"), "设置", self)
        setting_action.setShortcut("Ctrl+Alt+S")
        setting_action.triggered.connect(self.show_setting)

        exit_action = QAction(QIcon(r"exit.png"), "退出", self)
        exit_action.setShortcut("Ctrl+Alt+Q")
        exit_action.triggered.connect(qApp.quit)

        about_action = QAction(QIcon(r"info.png"), "关于", self)
        about_action.setShortcut("Ctrl+Alt+A")
        about_action.triggered.connect(self.show_about)

        file_menu.addAction(setting_action)
        file_menu.addAction(exit_action)
        help_menu.addAction(about_action)

        os.remove("setting.png")  # 删除临时图片
        os.remove("exit.png")
        os.remove("info.png")

    def init_layout(self):
        """
        设置布局
        :return: null
        """
        # 创建全局布局
        global_layout = QHBoxLayout()

        # 创建局部布局
        left_v_layout = QVBoxLayout()
        right_v_layout = QVBoxLayout()

        left_dir_h_layout = QHBoxLayout()
        left_file_h_layout = QHBoxLayout()
        left_apart_h_layout = QHBoxLayout()
        left_date_h_layout = QHBoxLayout()
        left_radio_h_layout = QHBoxLayout()

        left_dir_h_layout.addWidget(self.dir_label)
        left_dir_h_layout.addWidget(self.dir_edit)
        left_dir_h_layout.addWidget(self.dir_button)
        left_v_layout.addLayout(left_dir_h_layout)

        left_file_h_layout.addWidget(self.file_label)
        left_file_h_layout.addWidget(self.file_edit)
        left_file_h_layout.addWidget(self.file_button)
        left_v_layout.addLayout(left_file_h_layout)

        left_apart_h_layout.addWidget(self.apart_label)
        left_apart_h_layout.addWidget(self.apart_edit)
        left_apart_h_layout.addWidget(self.floor_edit)
        left_v_layout.addLayout(left_apart_h_layout)

        left_date_h_layout.addWidget(self.date_edit)
        left_date_h_layout.addWidget(self.time_edit)
        left_v_layout.addLayout(left_date_h_layout)

        left_radio_h_layout.addWidget(self.radio_button_hs)
        left_radio_h_layout.addWidget(self.radio_button_ky)
        left_v_layout.addLayout(left_radio_h_layout)

        left_v_layout.addWidget(self.push_button)
        left_v_layout.addWidget(self.process_output)

        right_v_layout.addWidget(self.table_widget)

        # 局部布局添加到全局布局
        global_layout.addLayout(left_v_layout)
        global_layout.addLayout(right_v_layout)
        global_layout.setStretch(0, 1)
        global_layout.setStretch(1, 2)

        # 一下三行用以解决使用了QMainWindow为父类时的布局问题
        widget = QWidget()
        widget.setLayout(global_layout)
        self.setCentralWidget(widget)

    def add_affairs(self):
        """
        添加消息响应
        :return: null
        """
        self.dir_edit.setFocusPolicy(Qt.NoFocus)
        self.file_edit.setFocusPolicy(Qt.NoFocus)
        self.date_edit.setFocusPolicy(Qt.NoFocus)
        self.process_output.setFocusPolicy(Qt.NoFocus)
        self.table_widget.setFocusPolicy(Qt.NoFocus)

        self.dir_button.clicked.connect(lambda: self.click_choice_dir())
        self.file_button.clicked.connect(lambda: self.click_choice_file())
        self.push_button.clicked.connect(lambda: self.click_start_pull())

        self.table_widget.setRowCount(32)
        self.table_widget.setColumnCount(3)
        self.table_widget.setHorizontalHeaderLabels(["宿舍号", "已收集数量", "操作"])
        self.table_widget.horizontalHeader().resizeSection(0, 194)
        self.table_widget.horizontalHeader().resizeSection(1, 194)
        self.table_widget.horizontalHeader().resizeSection(2, 194)
        self.table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 设置表格不可更改

    def setup(self):
        self.dir_edit.setText(Pb.dir_path)
        self.apart_edit.setText(Pb.apart)
        self.floor_edit.setCurrentIndex(Pb.floor_index)

    def click_choice_dir(self):
        """
        获取保存结果的路径
        :return: null
        """
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                             r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
        dir_path = QFileDialog.getExistingDirectory(self, "请选择文件夹路径", winreg.QueryValueEx(key, "Desktop")[0])
        self.dir_edit.setText(dir_path)

    def click_choice_file(self):
        """
        获取收集表关联的Excel表格文件
        :return: null
        """
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                             r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
        file, file_type = QFileDialog.getOpenFileName(self, "请选择文件", winreg.QueryValueEx(key, "Desktop")[0], "*.xlsx")
        self.file_edit.setText(file)

    def printf(self, p_str):
        """
        输出日志
        :param p_str: 日志内容
        :return: null
        """
        self.process_output.append(p_str)  # 在指定的区域显示提示信息
        self.cursor = self.process_output.textCursor()
        self.process_output.moveCursor(self.cursor.End)  # 光标移到最后，这样就会自动显示出来
        QtWidgets.QApplication.processEvents()  # 一定加上这个功能，不然有卡顿

    def pull_down(self, url, filename):
        """
        拉取材料
        :param url: 链接
        :param filename: 文件名
        :return:
        """
        if exists(filename) is False:
            try:
                urlretrieve(url=url, filename=filename)
            except urllib.error.URLError:
                QMessageBox.warning(self, "警告", "请检查是否开启了网络代理")
        if self.radio_button_hs.isChecked():
            text = pytesseract.image_to_string(filename, lang='chi_sim')
            # print(text)  # 输出OCR结果
            date_time = self.date_edit.text()
            test_text = date_time[0:4] + "-" + date_time[5:7] + "-" + date_time[8:10]
            if "采 样 时 间 : " + test_text in text or "采 样 时 间 " + test_text in text:
                self.printf(">> 截图符合要求")
            else:
                self.printf(">> 截图非当日检测结果")

    def click_start_pull(self):
        """
        拉取图片
        :return: null
        """
        if self.dir_edit.text() == "":
            QMessageBox.warning(self, '警告', '请选择保存路径')
        elif self.file_edit.text() == "":
            QMessageBox.warning(self, '警告', '请输选择文件路径')
        elif self.apart_edit.text() == "":
            QMessageBox.warning(self, '警告', '请输入楼栋编号')
        elif str(self.floor_edit.currentIndex()) == "0":
            QMessageBox.warning(self, '警告', '请选择楼层')
        else:
            res_path = self.dir_edit.text()
            apart = self.apart_edit.text()
            floor = self.floor_edit.currentText()
            date_time = self.date_edit.text()
            detail_time = "-" + self.time_edit.currentText() if str(self.time_edit.currentIndex()) != "0" else ""
            collect_type = "-核酸检测结果" if self.radio_button_hs.isChecked() else "-抗原检测结果"
            save_path = res_path + "/" + apart + "-" + floor + "-" + date_time + detail_time + collect_type

            self.printf(">> 准备拉取...\n")
            if exists(save_path) is False:
                makedirs(save_path)
            file = self.file_edit.text()
            sheet = load_sheet(file)
            dic = get_file_list(sheet)
            res_num = 0
            row = 0
            self.printf(">> 开始拉取...\n")
            for key in dic:
                folder = save_path + "/" + key
                if exists(folder) is False:
                    makedirs(folder)
                for items in dic[key]:
                    res_num += 1
                    url = items["img"]
                    path = folder + "/"
                    file_type = match(r'(.*)_type=([a-z]+)', url, M | I).group(2)
                    res_time = date_time[5:7] + date_time[8:10]
                    filename = path + items["id"] + "_" + items["name"] + "_" + res_time + "." + file_type
                    p = threading.Thread(target=self.pull_down, args=(url, filename))
                    p.start()
                p.join()  # 等待线程结束
                table_items = []
                nums = len(dic[key])
                table_items.append(key)
                table_items.append(nums)
                self.table_widget.setItem(row, 0, QTableWidgetItem(str(key)))
                self.table_widget.setItem(row, 1, QTableWidgetItem(str(nums)))
                open_button = QPushButton("打开文件夹")
                open_button.clicked.connect(
                    lambda: self.click_open_folder(save_path, self.table_widget.currentIndex().row()))
                self.table_widget.setCellWidget(row, 2, open_button)
                row += 1
                self.printf(">> " + key + "检测结果拉取完成")
            self.printf("\n>> 拉取结束，共 " + str(res_num) + " 份结果\n")

            if Pb.auto_pack == "true":
                self.printf(">> 开始自动打包\n")
                zip_file_name = save_path + ".zip"
                for turn in range(5):
                    shutil.make_archive(save_path, "zip", root_dir=save_path)
                    if (self.radio_button_ky.isChecked() and os.path.getsize(zip_file_name) < 10000000) or (
                            self.radio_button_hs.isChecked() and os.path.getsize(zip_file_name) < 1000000):
                        os.remove(zip_file_name)
                    else:
                        self.printf(">> 第 " + str(turn + 1) + " 次尝试打包成功\n")
                        self.printf(">> 自动打包完成\n")
                        break
                    if turn == 4:
                        self.printf("\n>> 自动打包失败! 请重试或手动打包!")
            else:
                self.printf(">> 已关闭自动打包，可前往'设置->高级'开启\n")

    def click_open_folder(self, path, row_index):
        """
        打开文件夹
        :param path: 路径
        :param row_index: 所在行数
        :return: null
        """
        folder = path + "/" + self.table_widget.item(row_index, 0).text()
        img_list = os.listdir(folder)
        read_list = []
        for img in img_list:
            image = cv2.imdecode(np.fromfile(folder + "/" + img, dtype=np.uint8), 1)
            if self.radio_button_ky.isChecked():
                h, w = image.shape[:2]  # 获取图像的高和宽
                blured = cv2.blur(image, (5, 5))  # 对图像进行均值滤波（低通滤波）来去掉噪声
                mask = np.zeros((h + 2, w + 2), np.uint8)  # 掩码长和宽都比输入图像多两个像素点，满水填充不会超出掩码的非零边缘
                cv2.floodFill(blured, mask, (w - 1, h - 1), (255, 255, 255), (2, 2, 2), (3, 3, 3),
                              8)  # 进行泛洪填充，类似PS中的魔棒工具，这里用于清除背景
                gray = cv2.cvtColor(blured, cv2.COLOR_BGR2GRAY)  # 得到灰度图
                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (100, 150))  # 定义结构元素
                opened = cv2.morphologyEx(gray, cv2.MORPH_OPEN, kernel)  # 开运算去除背景噪声
                closed = cv2.morphologyEx(opened, cv2.MORPH_CLOSE, kernel)  # 闭运算填充目标内的孔洞
                ret, binary = cv2.threshold(closed, 250, 255, cv2.THRESH_BINARY)  # 二值化
                contours, hierarchy = cv2.findContours(binary, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)  # 找到轮廓
                cv2.drawContours(image, contours, -1, (0, 0, 255), 10)  # 绘制轮廓
            # contour = find_contours(image)
            image = cv2.resize(image, dsize=(450, 800), interpolation=cv2.INTER_CUBIC)
            read_list.append(image)
        cv2.imshow("image", np.hstack(read_list))
        os.startfile(folder)

    def closeEvent(self, event):
        """
        关闭窗口事件
        :param event: 关闭事件
        :return: null
        """
        reply = QMessageBox.question(self, '提醒', '确定要退出吗?', QMessageBox.Yes, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.child_window.destroy()
            event.accept()
        else:
            event.ignore()

    def show_setting(self):
        Pb.has_printed = False
        self.child_window.setWindowModality(Qt.ApplicationModal)  # 阻塞主窗口
        self.child_window.show()
        self.child_window.signal.connect(self.get_data)

    def get_data(self, is_changed):
        if is_changed:
            self.setup()
            if not Pb.has_printed:
                self.printf(">> 修改设置成功\n")
                Pb.has_printed = True
            Pb.set_is_changed()

    def show_about(self):
        """
        关于框
        :return: null
        """
        QMessageBox.about(self, '关于', 'collector-v4.0\n豪杰收集助手\nCode By SHOU-1951124\nCode By SHOU-1951128')


if __name__ == "__main__":
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("collector")
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
